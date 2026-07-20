"""Exportadores: los archivos generados se abren y contienen lo esencial."""

import io

from openpyxl import load_workbook
from pptx import Presentation

from exportadores import excel, pptx

COLS = ["PAIS", "TOTAL_USD_FOB"]
FILAS = [["Estados Unidos", 4211591218.59], ["China", 987654321.0]]


def test_excel_se_abre_y_trae_datos() -> None:
    contenido = excel.construir("¿Top países?", "EE. UU. lidera.", "SELECT …", COLS, FILAS)
    wb = load_workbook(io.BytesIO(contenido))
    ws = wb.active
    plano = str(list(ws.iter_rows(values_only=True)))
    assert "Estados Unidos" in plano and "PAIS" in plano


def test_pptx_se_abre_con_laminas() -> None:
    contenido = pptx.construir("¿Top países?", "EE. UU. lidera.", "SELECT …", COLS, FILAS)
    prs = Presentation(io.BytesIO(contenido))
    assert len(prs.slides) >= 3
