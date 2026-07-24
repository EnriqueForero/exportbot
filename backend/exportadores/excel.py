"""Excel institucional de un resultado de ExportBot (openpyxl, server-side)."""

from __future__ import annotations

import io
from datetime import datetime
from zoneinfo import ZoneInfo

_TZ_BOGOTA = ZoneInfo("America/Bogota")
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import VERSION_APP
from exportadores import ADVERTENCIA_LEGAL, AZUL_INSTITUCIONAL

_MAX_ANCHO = 60


def construir(pregunta: str, texto: str, sql: str, columnas: list[str], filas: list[list[Any]]) -> bytes:
    """Construye el archivo .xlsx con pregunta, respuesta, tabla, SQL y advertencia."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    titulo = Font(bold=True, size=14, color=AZUL_INSTITUCIONAL)
    etiqueta = Font(bold=True, size=10, color=AZUL_INSTITUCIONAL)
    cab_fill = PatternFill("solid", fgColor=AZUL_INSTITUCIONAL)
    cab_font = Font(bold=True, color="FFFFFF")
    borde = Border(*(Side(style="thin", color="C9D2E0"),) * 4)

    ws["A1"] = "ExportBot 2.0 · ProColombia — Cifras de exportaciones de bienes"
    ws["A1"].font = titulo
    ws["A2"] = f"Generado: {datetime.now(_TZ_BOGOTA):%Y-%m-%d %H:%M} · Versión app: {VERSION_APP}"
    ws["A3"] = "Pregunta:"
    ws["A3"].font = etiqueta
    ws["B3"] = pregunta
    ws["A4"] = "Respuesta:"
    ws["A4"].font = etiqueta
    ws["B4"] = texto
    ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")

    fila_inicio = 6
    for j, col in enumerate(columnas, start=1):
        c = ws.cell(row=fila_inicio, column=j, value=col)
        c.fill, c.font, c.border = cab_fill, cab_font, borde
    for i, fila in enumerate(filas, start=fila_inicio + 1):
        for j, v in enumerate(fila, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = borde
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                c.number_format = "#,##0.00"
    for j, col in enumerate(columnas, start=1):
        ancho = max([len(str(col))] + [len(str(f[j - 1])) for f in filas[:200] if j - 1 < len(f)] + [10])
        ws.column_dimensions[get_column_letter(j)].width = min(ancho + 2, _MAX_ANCHO)

    fila_sql = fila_inicio + len(filas) + 2
    ws.cell(row=fila_sql, column=1, value="SQL ejecutada (auditoría):").font = etiqueta
    ws.cell(row=fila_sql + 1, column=1, value=sql).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=fila_sql + 3, column=1, value=ADVERTENCIA_LEGAL).alignment = Alignment(wrap_text=True)
    ws.cell(row=fila_sql + 3, column=1).font = Font(italic=True, size=8)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
